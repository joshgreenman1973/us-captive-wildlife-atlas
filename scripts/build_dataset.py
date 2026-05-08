"""
Build the combined facility dataset for the US Captive Wildlife Atlas.

Inputs:
  data/usda-licensees-raw.xlsx   — USDA APHIS active licensees & registrants
  data/aza-institutions.txt      — AZA-accredited US institutions (Name|City|ST)
  data/places_gazetteer.txt      — 2023 Census places gazetteer (city centroids)

Output:
  data/facilities.json           — combined, geocoded facility list

Geocoding precision is city-centroid only because USDA's public list does not
publish street addresses. This is documented on the methodology page.
"""

import csv
import json
import re
import unicodedata
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"

# License classes we keep on the map.
# Class C = Exhibitor (zoos, aquariums, roadside zoos, circuses).
# Class B = Dealer (animal dealers, often hold animals).
KEEP_CLASSES = {"Class C - Exhibitor", "Class B - Dealer"}

# Names that strongly suggest a domestic-pet operation rather than wildlife.
# Conservative — only flag Class B since most Class C "farm" names are
# legitimate wildlife (game farms, alligator farms, reindeer farms, etc.)
DOMESTIC_KW = re.compile(
    r"\b(pupp(?:y|ies)|kennel|cattery|pet shop|pet store|petland|petsmart"
    r"|rabbitry|hatchery|labradoodle|poodle|retriever|terrier|bulldog|spaniel"
    r"|shepherd|pomeranian|chihuahua|frenchie|maltese|yorkie"
    r"|cats? in the|cats? r us|dog house|dog haven|canine|kitten|puppymill)\b",
    re.I,
)

STATE_NAME_TO_ABBR = {
    "Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR",
    "California": "CA", "Colorado": "CO", "Connecticut": "CT", "Delaware": "DE",
    "District of Columbia": "DC", "Florida": "FL", "Georgia": "GA", "Hawaii": "HI",
    "Idaho": "ID", "Illinois": "IL", "Indiana": "IN", "Iowa": "IA",
    "Kansas": "KS", "Kentucky": "KY", "Louisiana": "LA", "Maine": "ME",
    "Maryland": "MD", "Massachusetts": "MA", "Michigan": "MI", "Minnesota": "MN",
    "Mississippi": "MS", "Missouri": "MO", "Montana": "MT", "Nebraska": "NE",
    "Nevada": "NV", "New Hampshire": "NH", "New Jersey": "NJ", "New Mexico": "NM",
    "New York": "NY", "North Carolina": "NC", "North Dakota": "ND", "Ohio": "OH",
    "Oklahoma": "OK", "Oregon": "OR", "Pennsylvania": "PA", "Rhode Island": "RI",
    "South Carolina": "SC", "South Dakota": "SD", "Tennessee": "TN", "Texas": "TX",
    "Utah": "UT", "Vermont": "VT", "Virginia": "VA", "Washington": "WA",
    "West Virginia": "WV", "Wisconsin": "WI", "Wyoming": "WY",
}


def norm_city(s: str) -> str:
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode()
    s = s.lower().strip()
    s = re.sub(r"[.,]", "", s)
    s = re.sub(r"-", " ", s)  # Winston-Salem -> winston salem
    s = re.sub(r"\s+", " ", s)
    # Strip common suffixes from gazetteer names
    s = re.sub(r"\s+(city|town|village|cdp|borough|township|municipality)$", "", s)
    # Common abbreviations in source data
    s = re.sub(r"^st ", "saint ", s)
    s = re.sub(r"^ft ", "fort ", s)
    s = re.sub(r"^mt ", "mount ", s)
    return s


# Manual overrides for cities the Census places file doesn't cover (NYC
# boroughs, common townships, unincorporated areas).
MANUAL_GEO = {
    ("brooklyn", "NY"): (40.6782, -73.9442),
    ("bronx", "NY"): (40.8448, -73.8648),
    ("queens", "NY"): (40.7282, -73.7949),
    ("staten island", "NY"): (40.5795, -74.1502),
    ("manhattan", "NY"): (40.7831, -73.9712),
    ("jackson", "NJ"): (40.1140, -74.3557),
    ("myakka city", "FL"): (27.3553, -82.1750),
    ("natural bridge", "VA"): (37.6303, -79.5439),
    ("honolulu", "HI"): (21.3099, -157.8581),
    ("canyon country", "CA"): (34.4144, -118.4517),
    ("windham", "ME"): (43.7864, -70.4267),
    ("washington", "MI"): (42.7222, -83.0083),
    ("caulfield", "MO"): (36.6300, -91.9100),
}


def load_gazetteer() -> dict:
    """Map (norm_city, state) -> (lat, lon)."""
    g = {}
    with open(DATA / "places_gazetteer.txt", encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter="\t")
        # Header has trailing whitespace; strip keys
        for row in reader:
            row = {k.strip(): (v.strip() if v else "") for k, v in row.items()}
            state = row["USPS"]
            name = row["NAME"]
            try:
                lat = float(row["INTPTLAT"])
                lon = float(row["INTPTLONG"])
            except (KeyError, ValueError):
                continue
            key = (norm_city(name), state)
            # Keep first hit (gazetteer is alphabetical by state then name); for
            # duplicates, prefer the one with larger land area
            try:
                aland = float(row.get("ALAND", 0) or 0)
            except ValueError:
                aland = 0
            if key not in g or aland > g[key][2]:
                g[key] = (lat, lon, aland)
    return {k: (v[0], v[1]) for k, v in g.items()}


def geocode(city: str, state: str, gz: dict):
    nc = norm_city(city)
    if (nc, state) in MANUAL_GEO:
        return MANUAL_GEO[(nc, state)]
    if (nc, state) in gz:
        return gz[(nc, state)]
    # Try common variants
    variants = [
        nc.replace("saint ", "st "),
        nc.replace("st ", "saint "),
        nc.replace("mount ", "mt "),
        nc.replace("mt ", "mount "),
    ]
    for v in variants:
        if (v, state) in gz:
            return gz[(v, state)]
    # Try fuzzy: any gazetteer name in same state that startswith
    for (gname, gstate), coord in gz.items():
        if gstate == state and gname.startswith(nc) and len(nc) >= 4:
            return coord
    return None


def load_usda():
    wb = openpyxl.load_workbook(DATA / "usda-licensees-raw.xlsx", read_only=True)
    ws = wb[wb.sheetnames[0]]
    out = []
    for row in ws.iter_rows(min_row=12, values_only=True):
        # cols: (None, License Type, APHIS#, Account, DBA, City, State, Exp, None)
        license_type = row[1]
        if not license_type or license_type not in KEEP_CLASSES:
            continue
        out.append({
            "source": "USDA",
            "license_type": license_type,
            "license_no": row[2] or "",
            "name": (row[3] or "").strip(),
            "dba": (row[4] or "").strip(),
            "city": (row[5] or "").strip(),
            "state": (row[6] or "").strip(),
            "expiration": row[7] or "",
        })
    return out


def load_aza():
    out = []
    with open(DATA / "aza-institutions.txt", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            parts = line.split("|")
            if len(parts) != 3:
                continue
            name, city, st = parts
            out.append({
                "source": "AZA",
                "name": name.strip(),
                "city": city.strip(),
                "state": st.strip(),
            })
    return out


def merge(usda, aza):
    """Merge AZA accreditation flag onto USDA records where they match by
    (city, state) + fuzzy name. Output unified facility list."""
    # Index USDA by (city_norm, state)
    by_loc = defaultdict(list)
    for u in usda:
        by_loc[(norm_city(u["city"]), u["state"])].append(u)

    facilities = []
    matched_usda_ids = set()

    for a in aza:
        loc_key = (norm_city(a["city"]), a["state"])
        candidates = by_loc.get(loc_key, [])
        # Try simple keyword match: any USDA record in same city whose name
        # shares 2+ tokens with the AZA name
        a_tokens = set(re.findall(r"[a-z]+", a["name"].lower())) - {
            "the", "of", "and", "at", "in", "on", "for", "zoo", "aquarium",
            "park", "society", "garden", "gardens", "inc", "llc"
        }
        match = None
        for c in candidates:
            c_tokens = set(re.findall(r"[a-z]+", c["name"].lower()))
            if len(a_tokens & c_tokens) >= 1:
                match = c
                break
        if match:
            matched_usda_ids.add(match["license_no"])
            facilities.append({
                "name": a["name"],
                "city": a["city"],
                "state": a["state"],
                "aza": True,
                "usda_class": match["license_type"],
                "usda_license": match["license_no"],
            })
        else:
            facilities.append({
                "name": a["name"],
                "city": a["city"],
                "state": a["state"],
                "aza": True,
                "usda_class": None,
                "usda_license": None,
            })

    # Add unmatched USDA records
    for u in usda:
        if u["license_no"] in matched_usda_ids:
            continue
        nm = u["dba"] if u["dba"] else u["name"]
        facilities.append({
            "name": nm,
            "city": u["city"],
            "state": u["state"],
            "aza": False,
            "usda_class": u["license_type"],
            "usda_license": u["license_no"],
        })
    return facilities


def main():
    print("Loading gazetteer...")
    gz = load_gazetteer()
    print(f"  {len(gz)} places")

    print("Loading USDA...")
    usda = load_usda()
    print(f"  {len(usda)} kept records (Class B + C)")

    print("Loading AZA...")
    aza = load_aza()
    print(f"  {len(aza)} AZA institutions")

    print("Merging...")
    facilities = merge(usda, aza)
    print(f"  {len(facilities)} unified facility records")

    print("Geocoding...")
    geocoded = 0
    missing = []
    for f in facilities:
        coord = geocode(f["city"], f["state"], gz)
        if coord:
            f["lat"], f["lon"] = coord
            geocoded += 1
        else:
            f["lat"], f["lon"] = None, None
            missing.append(f"{f['city']}, {f['state']}")
    print(f"  geocoded {geocoded} / {len(facilities)}")
    if missing:
        miss_counts = defaultdict(int)
        for m in missing:
            miss_counts[m] += 1
        print(f"  {len(set(missing))} unique unmatched cities; top 10:")
        for k, v in sorted(miss_counts.items(), key=lambda x: -x[1])[:10]:
            print(f"    {v}x  {k}")

    # Tag domestic-likely (Class B only) and apply deterministic jitter to
    # facilities that share an exact coordinate so they don't pile up.
    coord_groups = defaultdict(list)
    for f in facilities:
        if f["lat"] is not None:
            coord_groups[(f["lat"], f["lon"])].append(f)
    import math
    for (lat, lon), group in coord_groups.items():
        if len(group) == 1:
            continue
        # Spread on a small ring (~150 m radius). Deterministic by name.
        n = len(group)
        for i, f in enumerate(group):
            angle = 2 * math.pi * i / n
            # ~0.0015 deg ≈ 165 m; scales roughly fine at typical lat
            f["lat"] = round(lat + 0.0015 * math.cos(angle), 6)
            f["lon"] = round(lon + 0.0015 * math.sin(angle), 6)

    for f in facilities:
        # AZA-accredited facilities are never flagged as domestic
        if f.get("aza"):
            f["domestic_likely"] = False
            continue
        f["domestic_likely"] = bool(
            f.get("usda_class") in ("Class B - Dealer", "Class C - Exhibitor")
            and DOMESTIC_KW.search(f["name"])
        )

    # Filter to mappable
    mappable = [f for f in facilities if f["lat"] is not None]
    aza_count = sum(1 for f in mappable if f["aza"])
    usda_only = sum(1 for f in mappable if not f["aza"])

    out_path = DATA / "facilities.json"
    with open(out_path, "w") as f:
        json.dump({
            "generated": "2026-05-07",
            "source_dates": {
                "usda": "2026-04-20",
                "aza": "2026-05-07",
                "gazetteer": "2023",
            },
            "counts": {
                "total": len(facilities),
                "mapped": len(mappable),
                "aza_accredited": aza_count,
                "usda_only": usda_only,
            },
            "facilities": mappable,
        }, f, indent=1)
    print(f"Wrote {out_path}")
    print(f"  AZA-accredited: {aza_count}")
    print(f"  USDA-only: {usda_only}")


if __name__ == "__main__":
    main()
